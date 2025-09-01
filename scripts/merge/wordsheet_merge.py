import os
from tkinter import Tk
from tkinter.filedialog import askopenfilenames

from openpyxl import Workbook
from openpyxl.comments import Comment


def load_reference_file(reference_path):
    ref_wb = openpyxl.load_workbook(reference_path, read_only=True, keep_vba=True)

    def get_single_characters(sheet):
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        char_col_indices = [i for i, h in enumerate(headers) if h in ("單字", "单字")]

        if len(char_col_indices) == 0:
            raise ValueError("未找到“單字”或“单字”列。")
        if len(char_col_indices) > 1:
            raise ValueError("表中同时存在“單字”与“单字”列，请删除其中一个以避免歧义。")

        char_col_idx = char_col_indices[0]
        chars = []
        for row in sheet.iter_rows(min_row=2):
            value = row[char_col_idx].value
            if value:
                chars.append(str(value))
        return chars

    sheetnames = ref_wb.sheetnames
    main_sheet = None
    supplement_sheet = None

    if "主表" in sheetnames:
        main_sheet = ref_wb["主表"]
    else:
        main_sheet = ref_wb[sheetnames[0]]

    if "補充表" in sheetnames:
        supplement_sheet = ref_wb["補充表"]

    main_chars = get_single_characters(main_sheet)
    main_set = set(main_chars)

    if supplement_sheet:
        supplement_chars = get_single_characters(supplement_sheet)
        additional_chars = [char for char in supplement_chars if char not in main_set]
    else:
        additional_chars = []

    final_chars = main_chars.copy()
    if additional_chars:
        final_chars.append("-")
        final_chars.extend(additional_chars)

    return final_chars


# 处理多个用户选择的Excel表格
def select_excel_files():
    Tk().withdraw()  # 隐藏Tkinter的主窗口
    files = askopenfilenames(filetypes=[("Excel files", "*.xlsx;*.xlsm")])
    return files


# 合并字表
import openpyxl
from collections import defaultdict


def merge_excel_files(reference_chars, files):
    merged_data = {char: [""] * len(files) for char in reference_chars}  # 存放每个表的 syllable
    comments_data = {char: [[] for _ in range(len(files))] for char in reference_chars}  # 存放每个表的批注

    # 定义可接受的列名别名
    column_aliases = {
        'phrase': ['phrase', '單字', '单字'],
        'syllable': ['syllable', 'IPA', 'ipa'],
        'notes': ['notes', '注释', '注釋']
    }

    def find_column_index(header, targets):
        """在 header 中找出第一个匹配的 targets 项，并返回其索引"""
        for name in targets:
            if name in header:
                return header.index(name)
        return None

    for file_index, file in enumerate(files):
        wb = openpyxl.load_workbook(file, read_only=True)
        for ws in wb.worksheets:
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

            phrase_col = find_column_index(header, column_aliases['phrase'])
            syllable_col = find_column_index(header, column_aliases['syllable'])
            notes_col = find_column_index(header, column_aliases['notes'])

            # 跳过没有核心列的表
            if phrase_col is None or syllable_col is None:
                continue

            # 统计 phrase 重复数
            phrase_count = defaultdict(int)
            for row in ws.iter_rows(min_row=2):
                phrase = row[phrase_col].value
                phrase_count[phrase] += 1

            # 遍历数据行
            for row in ws.iter_rows(min_row=2):
                phrase = row[phrase_col].value
                syllable = row[syllable_col].value
                note = row[notes_col].value if notes_col is not None else None

                if phrase in merged_data:
                    if merged_data[phrase][file_index]:
                        merged_data[phrase][file_index] += f";{syllable}"
                        if note and phrase_count[phrase] > 1:
                            comments_data[phrase][file_index].append(note)
                    else:
                        merged_data[phrase][file_index] = syllable
                        if note and phrase_count[phrase] > 1:
                            comments_data[phrase][file_index].append(note)

    # 清理重复内容
    for phrase in merged_data:
        for i in range(len(merged_data[phrase])):
            entry = merged_data[phrase][i]
            if entry and ";" in entry:
                parts = [part.strip() for part in entry.split(";")]
                if all(p == parts[0] for p in parts):
                    merged_data[phrase][i] = parts[0]

    return merged_data, comments_data


# 获取文件名（不带路径和扩展名）
def get_file_name(file_path):
    return os.path.splitext(os.path.basename(file_path))[0]


# 创建新的Excel工作簿用于存储合并结果，并将批注加到syllable列的单元格上，保留原有的批注
def create_new_workbook(reference_chars, merged_data, comments_data, file_names):
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "字表"

    # 创建表头，第1列是'characters'，后续列根据文件名称生成
    headers = ['characters'] + file_names
    new_ws.append(headers)

    # 填充数据
    for row_idx, char in enumerate(reference_chars, start=2):  # 从第2行开始填充数据
        row_data = [char] + merged_data[char]
        new_ws.append(row_data)

        # 为每个文件列添加批注
        for col_idx in range(2, len(file_names) + 2):  # 从第2列开始到最后
            cell = new_ws.cell(row=row_idx, column=col_idx)
            file_index = col_idx - 2  # 计算在文件列表中的索引
            if comments_data[char][file_index]:
                comment_text = "; ".join(comments_data[char][file_index])
                if cell.comment:
                    # 如果已经有批注，保留原来的批注并追加新的内容
                    original_comment = cell.comment.text
                    new_comment_text = f"{original_comment}\n{comment_text}"
                    cell.comment = Comment(new_comment_text, "Python Script")
                else:
                    # 如果没有批注，直接添加新批注
                    cell.comment = Comment(comment_text, "Python Script")

    return new_wb


# 主函数
def merge_main():
    # 参考表路径
    reference_path = "参考表.xlsx"
    # reference_path = r"C:\Users\joengzaang\myfiles\杂文件\湛茂\茂名市信宜市金垌镇田心村_processed.xlsx"

    # 1. 读取参考表（只读模式）
    ref_chars = load_reference_file(reference_path)

    # 2. 选择多个Excel表
    selected_files = select_excel_files()

    if not selected_files:
        print("没有选择文件。")
        return

    # 获取文件名称作为列标题
    file_names = [get_file_name(file) for file in selected_files]

    # 3. 合并字表
    merged_data, comments_data = merge_excel_files(ref_chars, selected_files)

    # 4. 创建新表存储合并数据
    new_wb = create_new_workbook(ref_chars, merged_data, comments_data, file_names)

    # 5. 保存合并结果到新的Excel文件，文件名为"字表(总).xlsx"
    save_path = "merge.xlsx"
    new_wb.save(save_path)
    print(f"合并完成，结果已保存至 {save_path}")


if __name__ == "__main__":
    merge_main()
